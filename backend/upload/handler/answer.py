from uuid import uuid4

import openpyxl

from database import ProvidedReport, ImportAnswer, InitAnswer
from tools.date_time import DateTimeConvert
# from tools.dt import string_to_date, string_to_datetime, date_to_string, datetime_timezone


class AnswerChecker:

    def __init__(self, session, provided_report_id, import_filename, init_filename):
        self.session = session
        self.provided_report = session.query(ProvidedReport).filter(ProvidedReport.id == provided_report_id).scalar()
        self.messages = []
        self.import_filename = import_filename
        self.init_filename = init_filename
        self.answer = {
            'provided_report_id': provided_report_id,
            'import_data': None,
            'init_data': [],
            'is_ok': True
        }
        self.read_import()
        self.read_init()
        self.check()

    @property
    def result(self):
        return ', '.join(self.messages)

    def read_import(self):
        try:
            wb = openpyxl.load_workbook(self.import_filename)
            ws = wb.active
            self.answer['import_data'] = {
                'import_date': DateTimeConvert(ws['B3'].value).date,  # должна быть строка
                'reg_number': ws['B4'].value,
                'user': ws['B5'].value,
                'result': ws['B9'].value
            }
        except:
            self.messages.append('ошибка чтения протокола загрузки {}'.format(self.import_filename))
            self.answer['import_data'] = None

    def read_init(self):
        try:
            wb = openpyxl.load_workbook(self.init_filename)
            ws = wb.active
            for index, row in enumerate(ws.iter_rows()):
                if index >= 1 and row[0].value:
                    record = {
                        'number': int(row[0].value),
                        'fio': row[1].value.replace(u'\ufeff', '', 1),
                        'birthday': DateTimeConvert(row[2].value).date,  # должна быть строка
                        'init_date': DateTimeConvert(row[3].value).datetime,  # должно быть дата/время
                        'status': row[4].value if row[4].value else '',
                        'comment': row[5].value if row[5].value else '',
                        'provided_report_record_id': None
                    }
                    self.answer['init_data'].append(record)
        except:
            self.messages.append('ошибка чтения протокола идентификации {}'.format(self.init_filename))
            self.answer['init_data'] = []

    def check(self):
        if self.provided_report and self.answer['import_data'] and self.answer['init_data']:
            for record in self.answer['init_data']:
                if record['init_date'].date() < self.answer['import_data']['import_date']:
                    self.messages.append('дата импорта не соответствует дате идентификации')
                    self.answer['is_ok'] = False
                    return

            del_list = []
            for record in self.answer['init_data']:
                for item in self.provided_report.provided_report_records:
                    tmp = item.keeped_report_record.linked_defender.linked_person
                    if record['fio'] == tmp.person_appeal and DateTimeConvert(record['birthday']).string == tmp.birthday:  # должна быть дата
                        if item.id not in del_list:
                            record['provided_report_record_id'] = item.id
                            del_list.append(item.id)
                            break

            for record in self.answer['init_data']:
                if not record['provided_report_record_id']:
                    self.messages.append('{} ({}) отсутствует в отправленных записях'.format(record['fio'], record['birthday']))
                    self.answer['is_ok'] = False

            for item in self.provided_report.provided_report_records:
                print(item.keeped_report_record.linked_defender.linked_person.person_appeal, item.id)
                is_found = False
                for record in self.answer['init_data']:
                    if item.id == record['provided_report_record_id']:
                        is_found = True
                        print('found', record['provided_report_record_id'])
                if not is_found:
                    print('not found\n')
                    self.messages.append('{} ({}) отсутствует в протоколе идентификации'.format(
                        item.keeped_report_record.linked_defender.linked_person.person_appeal,
                        item.keeped_report_record.linked_defender.linked_person.birthday
                    ))
                    self.answer['is_ok'] = False
        else:
            if not self.provided_report:
                self.messages.append('зарегистрированные данные {} для СФР не найдены'.format(self.answer['provided_report_id']))
            self.answer['is_ok'] = False

    def save(self):
        answer_import = self.session.query(ImportAnswer).filter(ImportAnswer.provided_report_id == self.answer['provided_report_id']).scalar()
        if answer_import:
            self.messages.append('протокол импорта СФР на {} уже загружен'.format(self.answer['provided_report_id']))
            return
        else:
            answer_import = ImportAnswer()
            answer_import.id = str(uuid4())
            answer_import.created_utc = DateTimeConvert().value  # должно быть дата/время с часовым поясом
            answer_import.provided_report_id = self.answer['provided_report_id']
            answer_import.import_date = self.answer['import_data']['import_date']
            answer_import.reg_number = self.answer['import_data']['reg_number']
            answer_import.user = self.answer['import_data']['user']
            answer_import.result = self.answer['import_data']['result']
            self.session.add(answer_import)
            self.session.flush()

        for item in self.answer['init_data']:
            answer_init = self.session.query(InitAnswer).filter(InitAnswer.provided_report_record_id == item['provided_report_record_id']).scalar()
            if answer_init:
                self.messages.append('протокол идентификации СФР на {} уже загружен'.format(item['provided_report_record_id']))
                self.session.rollback()
                return
            else:
                answer_init = InitAnswer()
                answer_init.id = str(uuid4())
                answer_init.created_utc = DateTimeConvert().value  # должно быть дата/время с часовым поясом
                answer_init.provided_report_record_id = item['provided_report_record_id']
                answer_init.init_date = item['init_date']
                answer_init.number = item['number']
                answer_init.status = item['status']
                answer_init.comment = item['comment']
                self.session.add(answer_init)
                self.session.flush()
        self.messages.append('Успешно')
