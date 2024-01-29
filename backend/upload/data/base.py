from abc import ABC, abstractmethod
import openpyxl
import pyexcel

from tools import File


class BaseWorkbook(ABC):
    """
    Абстрактный класс загрузки и работы с книгами ODS, XLS, XLSX
    """

    def __init__(self, file_path):
        self.file = File(file_path)  # файл рабочей книги
        # self.file = file  # файл рабочей книги File()
        self.active = None  # имя активного листа рабочей книги
        self._data_ = None  # содержание загруженной рабочей книги
        if not self.file.is_exists:
            raise FileNotFoundError('файл <{}> не существует'.format(self.file.path))
        if self.file.extension not in ['.ods', '.xls', '.xlsx']:
            raise ImportError('недопустимое расширение файла <{}>'.format(self.file.extension))

    def __raiser__(self):
        """
        Создатель типовых исключений
        """
        if not self.file:
            raise ImportError('файл не выбран')
        if not self._data_:
            raise ImportError('содержимое файла не загружено')
        if not self.active:
            raise ImportError('активный лист не установлен')

    def load(self):
        """
        Загрузка содержания рабочей книги
        """
        try:
            if self.file.extension in ['.ods', '.xls']:
                self._data_ = pyexcel.get_book(file_name=self.file.path)
            elif self.file.extension == '.xlsx':
                self._data_ = openpyxl.load_workbook(self.file.path, read_only=True)
        except Exception as e:
            raise RuntimeError('содержимое не может быть прочитано <{}>'.format(e))

    @abstractmethod
    def check_active_worksheet(self):
        pass

    @abstractmethod
    def worksheet_data(self):
        pass

    def select_worksheet(self, name: str):
        """
        Выбор активного листа рабочей книги
        """
        if name in self.worksheets_list:
            self.active = name
        else:
            raise KeyError('лист <{}> не найден'.format(name))

    def get_row_num_data(self, row_num: int):
        """
        Данные строки в активном листе по её номеру
        """
        if self.file and self._data_ and self.active:
            try:
                if self.file.extension in ['.ods', '.xls']:
                    return self._data_.sheet_by_name(self.active).array[row_num - 1]
                elif self.file.extension == '.xlsx':
                    result = []
                    for cell in self._data_[self.active][row_num]:
                        result.append(cell.value)
                    return result
            except Exception as e:
                raise ValueError('строка #{} ошибка <{}>'.format(row_num, e))
        else:
            self.__raiser__()

    @property
    def worksheets_list(self):
        """
        Список листов в рабочей книге
        """
        result = []
        if self.file and self._data_:
            if self.file.extension in ['.ods', '.xls']:
                result = self._data_.sheet_names()
            elif self.file.extension == '.xlsx':
                for worksheet in self._data_.worksheets:
                    result.append(worksheet.title)
            return sorted(result)
        else:
            self.__raiser__()

    @property
    def max_row_num(self):
        """
        Максимальный номер строки в активном листе рабочей книги
        """
        if self.file and self._data_ and self.active:
            if self.file.extension in ['.ods', '.xls']:
                return len(self._data_.sheet_by_name(self.active).array)
            elif self.file.extension == '.xlsx':
                return self._data_[self.active].max_row
        else:
            self.__raiser__()
