from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill

from database import KeepedReport, KeepedReportRecord
from .data import result_file_data, protocol_import_data, protocol_init_data


def create_import_protocol(session, import_id, destination):
    # книга
    workbook = Workbook()
    worksheet = workbook.active
    # данные
    for row in protocol_import_data(session, import_id):
        worksheet.append(row)
    # определение ширины
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        adjusted_width = (max_length + 2) * 1.05
        worksheet.column_dimensions[column_letter].width = adjusted_width
    # оформление
    bd = Side(style='thin', color="000000")
    border = Border(left=bd, top=bd, right=bd, bottom=bd)
    header = NamedStyle(name="header")
    header.font = Font(name='Arial', bold=True)
    header.alignment = Alignment(horizontal='center', vertical='center')
    header.border = border
    content = NamedStyle(name="content")
    content.font = Font(name='Arial')
    content.alignment = Alignment(horizontal='left', vertical='top')
    content.border = border
    workbook.add_named_style(header)
    workbook.add_named_style(content)
    for cell in worksheet[1]:
        cell.style = header
    for row_num in range(2, worksheet.max_row + 1):
        for cell in worksheet[row_num]:
            cell.style = content
    # сохранить
    workbook.save(destination.path)


def create_init_protocol(session, import_id, destination):
    # книга
    workbook = Workbook()
    worksheet = workbook.active
    # данные
    for row in protocol_init_data(session, import_id):
        worksheet.append(row)
    # определение ширины
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        adjusted_width = (max_length + 2) * 1.05
        worksheet.column_dimensions[column_letter].width = adjusted_width
    worksheet.column_dimensions['F'].width = 50
    worksheet.column_dimensions['G'].width = 50
    # оформление
    bd = Side(style='thin', color="000000")
    border = Border(left=bd, top=bd, right=bd, bottom=bd)
    header = NamedStyle(name="header")
    header.font = Font(name='Arial', bold=True)
    header.alignment = Alignment(horizontal='center', vertical='center')
    header.border = border
    content = NamedStyle(name="content")
    content.font = Font(name='Arial')
    content.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    content.border = border
    workbook.add_named_style(header)
    workbook.add_named_style(content)
    for cell in worksheet[1]:
        cell.style = header
    for row_num in range(2, worksheet.max_row - 1):
        for cell in worksheet[row_num]:
            cell.style = content
    # сохранить
    workbook.save(destination.path)


def create_result_file(session, import_id, destination):
    # книга
    workbook = Workbook()
    worksheet = workbook.active

    # оформление
    bd = Side(style='thin', color="000000")
    border = Border(left=bd, top=bd, right=bd, bottom=bd)

    header = NamedStyle(name="header")
    header.font = Font(name='Arial', bold=True)
    header.alignment = Alignment(horizontal='center', vertical='center')
    header.border = border
    workbook.add_named_style(header)

    content = NamedStyle(name="content")
    content.font = Font(name='Arial')
    content.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    content.border = border
    workbook.add_named_style(content)

    warning_fill = NamedStyle(name="warning_fill")
    warning_fill.fill = PatternFill('solid', fgColor="FFFFBF")
    workbook.add_named_style(warning_fill)

    critical_fill = NamedStyle(name="critical_fill")
    critical_fill.font = Font(name='Arial')
    critical_fill.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    critical_fill.border = border
    critical_fill.fill = PatternFill('solid', fgColor="FF9C9C")
    workbook.add_named_style(critical_fill)

    # данные
    data = result_file_data(session, import_id)
    for index, row in enumerate(data):
        worksheet.append(row['data'])

        # for w in row['warnings']:
        #     if row['warnings'][w]:
        #         worksheet.cell(row=index + 1, column=int(w)+2).style = warning_fill

        if index < len(data) - 2:
            for cell in worksheet[index + 1]:
                cell.style = content

        for c in row['critical']:
            if row['critical'][c]:
                worksheet.cell(row=index + 1, column=int(c) + 2).style = critical_fill

    # определение ширины
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        adjusted_width = (max_length + 2) * 1.05
        worksheet.column_dimensions[column_letter].width = adjusted_width

    for cell in worksheet[1]:
        cell.style = header

    for cell in worksheet[2]:
        cell.style = header
    #
    # for row_num in range(3, worksheet.max_row - 1):
    #     for cell in worksheet[row_num]:
    #         cell.style = content

    # сохранить
    workbook.save(destination.path)


def create_sfr_control_file(header, rows, destination):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(header)
    for row in rows:
        worksheet.append(row)
    workbook.save(destination)


def create_sfr_forgiven_file(header, rows, destination):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(header)
    for row in rows:
        worksheet.append(row)
    workbook.save(destination)
