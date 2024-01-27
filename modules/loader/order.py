import openpyxl

# from tools import get_hash_md5, base_filename
from tools.date_and_time import datetime_timezone
from ._workbook_ import WorkBook


class OrdersWorkbook(WorkBook):
    """
    Класс загрузки приказов ОШУ Росгвардии об участии в СВО
    """

    def __init__(self, **keywords):
        super().__init__(**keywords)
        self.import_id = '{}'.format(datetime_timezone().strftime('%Y%m%d-%H%M%S'))

    def check(self):
        """
        Реализация абстрактного метода проверки структуры книги
        :return: bool
        """
        self._add_to_log_('Проверка: <{}>'.format(self.filename))
        required = ['СПИСОК', 'Моб', 'Территориальные органы']  # наименования листов с данными
        for item in required:
            if item not in self.worksheets_list:
                self._add_to_log_(
                    'Ошибка проверки: структура книги не соответствует заданному шаблону ({} not present in {})'.format(
                        item,
                        self.worksheets_list
                    )
                )
                return False
        return True

    def get_main_worksheet_data(self):
        """
        Запрос данных об участниках СВО из субъектов войск
        :return: [[], [], ...]
        """
        return self.get_worksheet_data('СПИСОК')

    def get_mob_worksheet_data(self):
        """
        Запрос данных о мобилизованных участниках СВО
        :return: [[], [], ...]
        """
        return self.get_worksheet_data('Моб')

    def get_terra_worksheet_data(self):
        """
        Запрос данных об участниках СВО из территориальных органов
        :return: [[], [], ...]
        """
        return self.get_worksheet_data('Территориальные органы')


class OrdersWorkbookPXL:

    def __init__(self, filepath):
        self._required_worksheets_ = ['СПИСОК', 'Моб', 'Территориальные органы']
        self.import_id = '{}'.format(datetime_timezone().strftime('%Y%m%d-%H%M%S'))
        self.filepath = filepath
        self.hashsum = get_hash_md5(self.filepath)
        self.workbook = None

    @property
    def filename(self):
        return base_filename(self.filepath)

    def load_workbook(self):
        try:
            self.workbook = openpyxl.load_workbook(self.filepath, read_only=True)
        except Exception as e:
            print(e)

    def _sheet_rows_(self, sheet_name):
        if self.workbook and sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            return None

    @property
    def main_sheet_rows(self):
        return self._sheet_rows_(self._required_worksheets_[0])

    @property
    def mob_sheet_rows(self):
        return self._sheet_rows_(self._required_worksheets_[1])

    @property
    def terra_sheet_rows(self):
        return self._sheet_rows_(self._required_worksheets_[2])

    def close_workbook(self):
        self.workbook.close()
